import os
import openpyxl
import csv
from io import BytesIO

def excel_to_csv(excel_path, csv_path):
    ext = os.path.splitext(excel_path)[1].lower()
    if ext == '.xlsx':
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["GradeDist"]

        # Set B1 to "Catalog Nbr"
        ws["B1"] = "Catalog Nbr"

        # Fill columns W, X, Y, Z, AA with space if empty
        for row in ws.iter_rows(min_row=2):
            for col in [23, 24, 25, 26, 27]:  # W=23, X=24, Y=25, Z=26, AA=27
                cell = row[col-1]
                if cell.value is None:
                    cell.value = " "

        # Save to a temporary CSV buffer
        temp_csv = BytesIO()
        ws_data = ws.values
        writer = csv.writer(temp_csv)
        for row in ws_data:
            writer.writerow(list(row))
        temp_csv.seek(0)
        lines = temp_csv.read().decode().splitlines()
    elif ext == '.xlsb':
        try:
            from pyxlsb import open_workbook
        except ImportError:
            raise ImportError("pyxlsb is required for .xlsb support. Install with 'pip install pyxlsb'.")
        lines = []
        with open_workbook(excel_path) as wb:
            ws = wb.get_sheet('GradeDist')
            # Build rows
            for i, row in enumerate(ws.rows()):
                values = [cell.v if cell.v is not None else "" for cell in row]
                # Set B1
                if i == 0 and len(values) > 1:
                    values[1] = "Catalog Nbr"
                # Fill W,X,Y,Z,AA (cols 22-26, 0-indexed)
                if i > 0:
                    for col in [22, 23, 24, 25, 26]:
                        if col < len(values) and (values[col] is None or values[col] == ""):
                            values[col] = " "
                line = ','.join(f'"{str(v)}"' if ',' in str(v) or '"' in str(v) else str(v) for v in values)
                lines.append(line)
    else:
        raise ValueError("Unsupported file type: " + ext)

    import json
    # Read desired column order from config
    config_path = os.path.join(os.path.dirname(__file__), "column_order.json")
    with open(config_path, "r") as cf:
        desired_order = json.load(cf)

    parsed_rows = [line.split(',') for line in lines]
    header = parsed_rows[0]
    # Map header names to indices
    header_map = {name.strip(): idx for idx, name in enumerate(header)}

    # Build reordered rows
    reordered_rows = []
    # Write header in desired order
    reordered_rows.append(desired_order)
    for row in parsed_rows[1:]:
        new_row = []
        for col in desired_order:
            idx = header_map.get(col, None)
            new_row.append(row[idx] if idx is not None and idx < len(row) else "")
        reordered_rows.append(new_row)

    # Write to output CSV
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        for row in reordered_rows:
            writer.writerow(row)
